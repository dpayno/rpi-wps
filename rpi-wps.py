import serial
import time
import csv
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from pathlib import Path
from collections import defaultdict

# ========================
# Configuration Constants
# ========================
UART_PORT = '/dev/ttyS0'      # UART port on Raspberry Pi
BAUDRATE = 9600               # Communication speed
SAVE_INTERVAL = 10            # Seconds between writes
FILENAME_PREFIX = "pressure"  # Base name for log files

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ========================
# Utility Functions
# ========================
def is_float(value: str) -> bool:
    """Check if a string can be converted to a float."""
    try:
        float(value)
        return True
    except ValueError:
        return False

def get_usb_mount_point() -> Path | None:
    """Return the first USB mount point under /media/pi/, or None if not found."""
    base = Path("/media/pi")
    if base.exists() and base.is_dir():
        for sub in base.iterdir():
            if sub.is_dir():
                return sub
    return None

def get_daily_filename(extension: str):
    today_str = time.strftime("%Y-%m-%d")
    filename = f"{FILENAME_PREFIX}_{today_str}.{extension}"
    return filename

def get_output_paths(extension: str) -> list[Path]:
    """Generate a filename based on today's date and given extension, in local path."""
    filename = get_daily_filename(extension)
    paths = [str(Path.cwd() / filename)]
    usb_mount = get_usb_mount_point()
    if usb_mount:
        paths.append(str(usb_mount / filename))
    return paths

def initialize_excel(file_path: str):
    """Create a new Excel file with headers and empty chart area."""
    if not Path(file_path).exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # Main raw data table headers
        ws.append(["Timestamp", "Pressure (bar)"])
        # Minute averages header start at col D (4)
        ws.cell(row=1, column=4, value="Minute")
        ws.cell(row=1, column=5, value="Average Pressure (bar)")
        
        wb.save(file_path)

def load_excel(file_path: str):
    """Load an existing Excel file and return workbook and active sheet."""
    wb = load_workbook(file_path)
    ws = wb.active
    return wb, ws

def initialize_csv(file_path: str):
    """Create a new CSV file with headers if it doesn't exist."""
    if not Path(file_path).exists():
        with open(file_path, mode='w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Timestamp", "Pressure (bar)"])

def append_csv(file_path: str, rows: list):
    """Append multiple rows to a CSV file."""
    with open(file_path, mode='a', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)

def update_minute_averages_table(ws):
    """
    Calculate minute averages from raw data and write them to columns D and E.
    Format for minute: YYYY-MM-DD HH:MM (no seconds)
    """
    # Read all raw timestamps and pressures from columns A and B
    raw_data = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        raw_data.append(row)
    
    if not raw_data:
        return
    
    # Group values by minute
    grouped = defaultdict(list)
    for ts_str, pressure in raw_data:
        # Truncate seconds: keep YYYY-MM-DD HH:MM
        minute_str = ts_str[:16]
        grouped[minute_str].append(pressure)
    
    # Clear previous minute averages (rows 2..end in cols D and E)
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=5, value=None)
    
    # Write new minute averages starting from row 2
    row_idx = 2
    for minute, pressures in sorted(grouped.items()):
        avg_pressure = sum(pressures) / len(pressures)
        ws.cell(row=row_idx, column=4, value=minute)
        ws.cell(row=row_idx, column=5, value=round(avg_pressure, 3))
        row_idx += 1
    
    # Remove any charts before adding new one
    ws._charts.clear()
    
    # Create chart using minute averages
    chart = LineChart()
    chart.title = "Average Pressure per Minute"
    chart.x_axis.title = "Minute"
    chart.y_axis.title = "Pressure (bar)"
    
    data = Reference(ws, min_col=5, min_row=1, max_row=row_idx-1)  # avg pressure
    cats = Reference(ws, min_col=4, min_row=2, max_row=row_idx-1)  # minute labels
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G2")  # Put chart starting at G2

def read_sensor_line(serial_conn: serial.Serial) -> str:
    """Read a single line from the UART connection."""
    return serial_conn.readline().decode('utf-8').strip()

def process_sensor_data(line: str):
    """
    Validate and parse sensor data.
    Expected format: P=<value>
    Returns tuple (timestamp, float_value) if valid, otherwise None.
    """
    if line.startswith("P="):
        value_str = line[2:]
        if is_float(value_str):
            value = float(value_str)
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            return timestamp, value
    return None


# ========================
# Main Function
# ========================
def main():
    logger.info("Starting UART sensor logger with minute-averaged chart...")
    logger.info("Press Ctrl+C to stop.")

    current_date = time.strftime("%Y-%m-%d")
    excel_paths = get_output_paths("xlsx")
    csv_paths = get_output_paths("csv")
    for path in excel_paths:
        initialize_excel(path)
    for path in csv_paths:
        initialize_csv(path)
    wb, ws = load_excel(excel_paths[0])

    ser = serial.Serial(UART_PORT, baudrate=BAUDRATE, timeout=1)

    readings_buffer = []
    last_save_time = time.time()

    try:
        while True:
            new_date = time.strftime("%Y-%m-%d")

            # Rotate files daily
            if new_date != current_date:
                logger.info(f"Switching to new daily logs: {new_date}")

                if readings_buffer:
                    for r in readings_buffer:
                        ws.append(r)
                    update_minute_averages_table(ws)
                    for path in excel_paths:
                        wb.save(path)
                    for path in csv_paths:
                        append_csv(path, readings_buffer)
                    readings_buffer.clear()
                
                current_date = new_date
                excel_paths = get_output_paths("xlsx")
                csv_paths = get_output_paths("csv")
                for path in excel_paths:
                    initialize_excel(path)
                for path in csv_paths:
                    initialize_csv(path)
                wb, ws = load_excel(excel_paths[0])

            line = read_sensor_line(ser)
            if line:
                data = process_sensor_data(line)
                if data:
                    timestamp, value = data
                    logger.info(f"{timestamp} -> {value:.2f} bar")
                    readings_buffer.append(data)
                else:
                    logger.warning(f"Invalid data: {line}")

            if time.time() - last_save_time >= SAVE_INTERVAL and readings_buffer:
                for r in readings_buffer:
                    ws.append(r)
                update_minute_averages_table(ws)
                for path in excel_paths:
                    wb.save(path)
                for path in csv_paths:
                    append_csv(path, readings_buffer)
                logger.info(f"Saved {len(readings_buffer)} readings and updated chart.")
                readings_buffer.clear()
                last_save_time = time.time()

            time.sleep(0.1)

    except KeyboardInterrupt:
        logger.info("\nStopping monitoring and saving data...")
    except Exception as e:
        logger.exception(f"Unexpected error")
    finally:
        if readings_buffer:
            for r in readings_buffer:
                ws.append(r)
            update_minute_averages_table(ws)
            for path in excel_paths:
                wb.save(path)
            for path in csv_paths:
                append_csv(path, readings_buffer)
            logger.info(f"Saved {len(readings_buffer)} readings before exit and updated chart.")
        ser.close()


if __name__ == "__main__":
    main()

